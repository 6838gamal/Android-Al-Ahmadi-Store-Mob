import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from fastapi import FastAPI, Request, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
from typing import Optional
from types import SimpleNamespace
from datetime import datetime
import asyncio
import httpx
from urllib.parse import quote as _q
from contextlib import asynccontextmanager

# ── Backend API URL — يُقرأ من BACKEND_API_URL أو يستخدم القيمة الافتراضية ────
API_BASE = os.getenv("BACKEND_API_URL", "http://127.0.0.1:8000").rstrip("/")


@asynccontextmanager
async def _lifespan(app: FastAPI):
    yield


app = FastAPI(title="لوحة إدارة اندرويد الاحمدي", docs_url=None, redoc_url=None, lifespan=_lifespan)

@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    import traceback, html as _html
    tb = "".join(traceback.format_exception(type(exc), exc, exc.__traceback__))
    print(f"[UNHANDLED 500] {request.method} {request.url}\n{tb}")
    tb_escaped = _html.escape(tb)
    method_escaped = _html.escape(str(request.method))
    url_escaped = _html.escape(str(request.url))
    from fastapi.responses import HTMLResponse
    return HTMLResponse(
        f"""<!DOCTYPE html>
<html lang="ar" dir="rtl">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>خطأ في الخادم</title>
  <style>
    body {{ font-family: 'Segoe UI', sans-serif; background:#0D1117; color:#e6edf3; margin:0; padding:32px; }}
    .card {{ background:#161B22; border:1px solid #30363d; border-radius:12px; padding:28px; max-width:900px; margin:auto; }}
    h2 {{ color:#f85149; margin:0 0 8px; font-size:20px; }}
    .meta {{ color:#8b949e; font-size:13px; margin-bottom:20px; }}
    .trace {{ background:#0D1117; border:1px solid #30363d; border-radius:8px; padding:16px;
              font-family:monospace; font-size:12px; color:#e6edf3; white-space:pre-wrap;
              word-break:break-word; max-height:400px; overflow-y:auto; }}
    .back {{ display:inline-block; margin-top:20px; padding:10px 20px; background:#1A73E8;
             color:#fff; border-radius:8px; text-decoration:none; font-size:14px; }}
  </style>
</head>
<body>
  <div class="card">
    <h2>⚠️ حدث خطأ داخلي في الخادم</h2>
    <p class="meta">{method_escaped} {url_escaped}</p>
    <div class="trace">{tb_escaped}</div>
    <a href="javascript:history.back()" class="back">← العودة للصفحة السابقة</a>
  </div>
</body>
</html>""",
        status_code=500,
    )

app.add_middleware(
    SessionMiddleware,
    secret_key=os.getenv("SECRET_KEY", "android-alahmadi-replit-secret-key-2026-very-secure-random-string"),
    max_age=86400,
    https_only=False,
    same_site="lax",
)

UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")
app.mount("/static", StaticFiles(directory=os.path.join(os.path.dirname(__file__), "static")), name="static")

templates = Jinja2Templates(directory=os.path.join(os.path.dirname(__file__), "templates"))

# Safe enum/string value: {{ item.status|v }} works for both plain str and enum-like object
templates.env.filters['v'] = lambda x: (getattr(x, 'value', x) if x is not None else '')

# ── Helpers ────────────────────────────────────────────────────────────────────

_DT_KEYS = frozenset({
    "created_at", "updated_at", "expires_at", "starts_at", "ends_at",
    "responded_at", "warranty_start", "warranty_end", "resolved_at",
})

def _parse_dt(v: str):
    try:
        return datetime.fromisoformat(v.replace("Z", "+00:00")).replace(tzinfo=None)
    except Exception:
        return v

def to_obj(data):
    """Recursively convert API dicts/lists to attribute-accessible SimpleNamespace objects."""
    if isinstance(data, list):
        return [to_obj(i) for i in data]
    if isinstance(data, dict):
        converted = {}
        for k, v in data.items():
            if isinstance(v, str) and k in _DT_KEYS:
                converted[k] = _parse_dt(v)
            else:
                converted[k] = to_obj(v)
        return SimpleNamespace(**converted)
    return data


_API_TIMEOUT = 10.0  # Keep well below Replit proxy timeout (~30s)


async def api(method: str, path: str, token: str = None, real_ip: str = None, **kwargs):
    """Call the backend API; returns parsed JSON on success, None on error."""
    headers = {}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    if real_ip:
        headers["X-Real-IP"] = real_ip
        headers["X-Forwarded-For"] = real_ip
    try:
        async with httpx.AsyncClient(timeout=_API_TIMEOUT, follow_redirects=True) as client:
            resp = await getattr(client, method)(
                f"{API_BASE}{path}", headers=headers, **kwargs
            )
        if resp.status_code in (200, 201):
            return resp.json()
    except Exception as e:
        print(f"[admin] API error {method.upper()} {path}: {e}")
    return None


async def api_raw_upload(path: str, files: dict, data: dict = None, token: str = None):
    """Upload multipart/form-data to the backend. Returns (result, error_msg)."""
    headers = {}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    try:
        async with httpx.AsyncClient(timeout=60, follow_redirects=True) as client:
            resp = await client.post(
                f"{API_BASE}{path}", headers=headers, files=files, data=data or {}
            )
        if resp.status_code in (200, 201):
            return resp.json(), None
        try:
            body = resp.json()
            detail = body.get("detail", str(body)) if isinstance(body, dict) else str(body)
        except Exception:
            detail = resp.text[:300]
        return None, f"خطأ {resp.status_code}: {detail}"
    except Exception as e:
        return None, f"خطأ في الاتصال: {str(e)[:120]}"


async def api_ex(method: str, path: str, token: str = None, **kwargs):
    """Extended API call — returns (data, error_msg). data=None means failure."""
    headers = {}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    try:
        async with httpx.AsyncClient(timeout=_API_TIMEOUT, follow_redirects=True) as client:
            resp = await getattr(client, method)(
                f"{API_BASE}{path}", headers=headers, **kwargs
            )
        if resp.status_code in (200, 201):
            return resp.json(), None
        try:
            body = resp.json()
            if isinstance(body, dict):
                detail = body.get("detail", str(body))
                if isinstance(detail, list):
                    detail = "; ".join(
                        str(d.get("msg", d)) if isinstance(d, dict) else str(d)
                        for d in detail
                    )
            else:
                detail = str(body)
        except Exception:
            detail = resp.text[:300]
        print(f"[admin] {method.upper()} {path} → {resp.status_code}: {detail}")
        return None, f"خطأ {resp.status_code}: {detail}"
    except Exception as e:
        print(f"[admin] API error {method.upper()} {path}: {e}")
        return None, f"خطأ في الاتصال بالخادم: {str(e)[:120]}"


def _token(req: Request) -> str:
    return req.session.get("token", "")

def _name(req: Request) -> str:
    return req.session.get("admin_name", "المدير")

def _logged(req: Request) -> bool:
    return bool(req.session.get("admin_id"))

def _redirect_login():
    return RedirectResponse("/login", status_code=302)

# ── API Status Check (used by login page JS) ────────────────────────────────────

@app.get("/api-status")
async def api_status():
    """Check if backend API is reachable — called by login page JS before showing the form."""
    from fastapi.responses import JSONResponse
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            resp = await client.get(f"{API_BASE}/api/health")
        if resp.status_code == 200:
            return JSONResponse({"ok": True})
        return JSONResponse({"ok": False, "reason": f"HTTP {resp.status_code}"}, status_code=200)
    except Exception as e:
        return JSONResponse({"ok": False, "reason": str(e)[:120]}, status_code=200)


# ── Auth ────────────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def root(request: Request):
    return RedirectResponse("/dashboard" if _logged(request) else "/login", status_code=302)


@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    if _logged(request):
        return RedirectResponse("/dashboard", status_code=302)
    return templates.TemplateResponse(request, "login.html", {"error": None})


def _js_redirect(dest: str) -> HTMLResponse:
    """Return a 200 HTML page that immediately navigates to dest via JS.
    Avoids 303 redirect chains which can lose cookies in some proxy setups."""
    return HTMLResponse(f"""<!doctype html>
<html lang="ar" dir="rtl">
<head>
  <meta charset="UTF-8">
  <meta http-equiv="refresh" content="0;url={dest}">
  <script>window.location.replace({dest!r})</script>
  <style>body{{background:#0D1117;color:#e6edf3;font-family:sans-serif;
    display:flex;align-items:center;justify-content:center;height:100vh;margin:0}}</style>
</head>
<body><p>جارٍ التحويل…</p></body>
</html>""")


@app.post("/login", response_class=HTMLResponse)
async def login_post(request: Request, identifier: str = Form(...), password: str = Form(...)):
    identifier = identifier.strip()
    client_ip = (
        request.headers.get("x-real-ip")
        or request.headers.get("x-forwarded-for", "").split(",")[0].strip()
        or (request.client.host if request.client else "unknown")
    )
    async with httpx.AsyncClient(timeout=_API_TIMEOUT, follow_redirects=True) as client:
        try:
            headers_req = {"X-Real-IP": client_ip, "X-Forwarded-For": client_ip}
            resp = await client.post(
                f"{API_BASE}/api/auth/admin-login",
                json={"identifier": identifier, "password": password},
                headers=headers_req,
            )
            print(f"[login] status={resp.status_code} body={resp.text[:200]}")
            if resp.status_code == 429:
                return templates.TemplateResponse(request, "login.html",
                    {"error": "محاولات كثيرة. انتظر دقيقة ثم أعد المحاولة."})
            if resp.status_code in (200, 201):
                data = resp.json()
                user_data = data.get("user", {})
                if user_data.get("role") in ("admin", "branch_manager"):
                    request.session["admin_id"]   = user_data.get("id")
                    request.session["admin_name"] = user_data.get("name", "المدير")
                    request.session["token"]      = data.get("access_token")
                    # Return 200 + JS redirect — avoids 303 chain losing cookie in proxy
                    # If request came from Node proxy (127.0.0.1) → prefix with /admin-panel
                    via_proxy = request.client and request.client.host == "127.0.0.1"
                    dest = "/admin-panel/dashboard" if via_proxy else "/dashboard"
                    return _js_redirect(dest)
        except httpx.TimeoutException:
            print(f"[login] TIMEOUT connecting to API")
            return templates.TemplateResponse(request, "login.html",
                {"error": "⏳ الخادم في وضع الاستعداد ويستيقظ الآن — أعد المحاولة خلال 30 ثانية."})
        except Exception as e:
            print(f"[login] exception type={type(e).__name__} msg={e!r}")
            return templates.TemplateResponse(request, "login.html",
                {"error": "تعذّر الاتصال بالخادم. تحقق من اتصالك وأعد المحاولة."})
    return templates.TemplateResponse(request, "login.html",
                                      {"error": "بيانات الدخول غير صحيحة. تحقق من البريد وكلمة المرور."})


@app.post("/dashboard", response_class=HTMLResponse)
async def dashboard_post(request: Request):
    """Catch accidental POST to /dashboard (e.g. proxy redirect re-POST) → redirect to GET."""
    return RedirectResponse("/dashboard", status_code=303)


@app.get("/logout")
async def logout(request: Request):
    request.session.clear()
    return RedirectResponse("/login", status_code=302)


# ── Dashboard ───────────────────────────────────────────────────────────────────

@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard(request: Request):
    if not _logged(request):
        return _redirect_login()

    stats = await api("get", "/api/dashboard/stats", token=_token(request)) or {}
    recent_orders = [to_obj(o) for o in stats.get("recent_orders", [])]

    status_labels = {
        "received": "مستلم", "reviewing": "قيد المراجعة", "confirmed": "مؤكد",
        "preparing": "جاري التحضير", "shipped": "تم الشحن", "on_the_way": "في الطريق",
        "delivered": "تم التسليم", "cancelled": "ملغي",
    }
    status_colors = {
        "received": "#6B7280", "reviewing": "#F59E0B", "confirmed": "#3B82F6",
        "preparing": "#8B5CF6", "shipped": "#06B6D4", "on_the_way": "#10B981",
        "delivered": "#22C55E", "cancelled": "#EF4444",
    }

    return templates.TemplateResponse(request, "dashboard.html", {
        "admin_name":        _name(request),
        "active":            "dashboard",
        "total_orders":      stats.get("total_orders", 0),
        "active_orders":     stats.get("active_orders", 0),
        "total_revenue":     stats.get("total_revenue", 0),
        "total_products":    stats.get("total_products", 0),
        "total_customers":   stats.get("total_customers", 0),
        "maintenance_count": stats.get("maintenance_count", 0),
        "low_stock":         stats.get("low_stock_count", 0),
        "available_products":stats.get("available_products", 0),
        "recent_orders":     recent_orders,
        "status_labels":     status_labels,
        "status_colors":     status_colors,
    })


# ── Products ────────────────────────────────────────────────────────────────────

@app.get("/products", response_class=HTMLResponse)
async def products_list(request: Request, search: str = ""):
    if not _logged(request):
        return _redirect_login()
    params = {"limit": 500}
    if search:
        params["search"] = search
    raw = await api("get", "/api/products/", token=_token(request), params=params) or []
    products = to_obj(raw)
    status_map = {"available": "متوفر", "reserved": "محجوز", "sold": "مباع", "unavailable": "غير متوفر"}
    cat_map    = {"screen": "شاشة", "battery": "بطارية", "camera": "كاميرا", "speaker": "سماعة",
                  "charger": "شاحن", "device": "جهاز", "spare_part": "قطعة غيار", "other": "أخرى"}
    return templates.TemplateResponse(request, "products.html", {
        "admin_name": _name(request), "active": "products",
        "products": products, "search": search,
        "status_map": status_map, "cat_map": cat_map,
    })


_CATEGORIES = [
    ("screen","شاشة"), ("battery","بطارية"), ("camera","كاميرا"),
    ("speaker","سماعة"), ("charger","شاحن"), ("device","جهاز"),
    ("spare_part","قطعة غيار"), ("other","أخرى"),
]

from backend.core.samsung_catalog import SAMSUNG_CATALOG as _SAMSUNG_CATALOG

_SAMSUNG_SERIES = [
    {"key": k, "label_ar": v["label_ar"], "models": v["models"]}
    for k, v in _SAMSUNG_CATALOG.items()
]


@app.get("/products/add", response_class=HTMLResponse)
async def product_add_page(request: Request):
    if not _logged(request):
        return _redirect_login()
    return templates.TemplateResponse(request, "product_form.html", {
        "admin_name": _name(request), "product": None, "error": None,
        "categories": _CATEGORIES, "samsung_series": _SAMSUNG_SERIES,
    })


@app.post("/products/add", response_class=HTMLResponse)
async def product_add_post(
    request: Request,
    name: str = Form(...), name_ar: str = Form(""), brand: str = Form(""),
    model: str = Form(""), series: str = Form(""), category: str = Form(...),
    price: float = Form(...), quantity: int = Form(0),
    description: str = Form(""), notes: str = Form(""),
    is_featured: Optional[str] = Form(None),
    image: Optional[UploadFile] = File(None),
):
    if not _logged(request):
        return _redirect_login()

    image_url = None
    if image and image.filename:
        img_data = await api(
            "post", "/api/uploads/image", token=_token(request),
            files={"file": (image.filename, await image.read(), image.content_type)},
        )
        if img_data:
            image_url = img_data.get("url")

    payload = {
        "name": name, "name_ar": name_ar or None, "brand": brand or None,
        "model": model or None, "series": series or None,
        "category": category, "price": price,
        "quantity": quantity, "description": description or None,
        "notes": notes or None, "is_featured": bool(is_featured),
        "image_url": image_url,
    }
    _, err = await api_ex("post", "/api/products/", token=_token(request), json=payload)
    if err:
        return RedirectResponse(f"/products?error={_q(err)}", status_code=302)
    return RedirectResponse("/products?success=تم+إضافة+المنتج+بنجاح", status_code=302)


@app.get("/products/edit/{product_id}", response_class=HTMLResponse)
async def product_edit_page(product_id: int, request: Request):
    if not _logged(request):
        return _redirect_login()
    raw = await api("get", f"/api/products/{product_id}", token=_token(request))
    if not raw:
        return RedirectResponse("/products", status_code=302)
    return templates.TemplateResponse(request, "product_form.html", {
        "admin_name": _name(request), "product": to_obj(raw), "error": None,
        "categories": _CATEGORIES, "samsung_series": _SAMSUNG_SERIES,
    })


@app.post("/products/edit/{product_id}", response_class=HTMLResponse)
async def product_edit_post(
    product_id: int, request: Request,
    name: str = Form(...), name_ar: str = Form(""), brand: str = Form(""),
    model: str = Form(""), series: str = Form(""), category: str = Form(...),
    price: float = Form(...), quantity: int = Form(0),
    description: str = Form(""), notes: str = Form(""),
    is_featured: Optional[str] = Form(None), status: str = Form("available"),
    image: Optional[UploadFile] = File(None),
):
    if not _logged(request):
        return _redirect_login()

    image_url = None
    if image and image.filename:
        img_data = await api(
            "post", "/api/uploads/image", token=_token(request),
            files={"file": (image.filename, await image.read(), image.content_type)},
        )
        if img_data:
            image_url = img_data.get("url")

    payload = {
        "name": name, "name_ar": name_ar or None, "brand": brand or None,
        "model": model or None, "series": series or None,
        "category": category, "price": price,
        "quantity": quantity, "description": description or None,
        "notes": notes or None, "is_featured": bool(is_featured), "status": status,
    }
    if image_url:
        payload["image_url"] = image_url
    _, err = await api_ex("put", f"/api/products/{product_id}", token=_token(request), json=payload)
    if err:
        return RedirectResponse(f"/products?error={_q(err)}", status_code=302)
    return RedirectResponse("/products?success=تم+تحديث+المنتج+بنجاح", status_code=302)


@app.post("/products/delete/{product_id}")
async def product_delete(product_id: int, request: Request):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("delete", f"/api/products/{product_id}", token=_token(request))
    if err:
        return RedirectResponse(f"/products?error={_q(err)}", status_code=302)
    return RedirectResponse("/products?success=تم+حذف+المنتج+بنجاح", status_code=302)


# ── Orders ──────────────────────────────────────────────────────────────────────

@app.get("/orders", response_class=HTMLResponse)
async def orders_list(request: Request, status_filter: str = "", search: str = ""):
    if not _logged(request):
        return _redirect_login()
    params = {"limit": 300, "order_type": "product"}
    if status_filter:
        params["status"] = status_filter
    if search:
        params["search"] = search
    raw = await api("get", "/api/orders/", token=_token(request), params=params) or []
    orders = to_obj(raw)
    status_labels = {
        "received": "مستلم", "reviewing": "قيد المراجعة", "confirmed": "مؤكد",
        "preparing": "جاري التحضير", "shipped": "تم الشحن", "on_the_way": "في الطريق",
        "delivered": "تم التسليم", "cancelled": "ملغي",
    }
    status_colors = {
        "received": "#6B7280", "reviewing": "#F59E0B", "confirmed": "#3B82F6",
        "preparing": "#8B5CF6", "shipped": "#06B6D4", "on_the_way": "#10B981",
        "delivered": "#22C55E", "cancelled": "#EF4444",
    }
    return templates.TemplateResponse(request, "orders.html", {
        "admin_name": _name(request), "active": "orders",
        "orders": orders, "status_labels": status_labels, "status_colors": status_colors,
        "status_filter": status_filter, "search": search,
        "all_statuses": list(status_labels.items()),
    })


@app.get("/orders/{order_id}", response_class=HTMLResponse)
async def order_detail(order_id: int, request: Request):
    if not _logged(request):
        return _redirect_login()
    raw = await api("get", f"/api/orders/{order_id}", token=_token(request))
    if not raw:
        return RedirectResponse("/orders", status_code=302)
    status_labels = {
        "received": "مستلم", "reviewing": "قيد المراجعة", "confirmed": "مؤكد",
        "preparing": "جاري التحضير", "shipped": "تم الشحن", "on_the_way": "في الطريق",
        "delivered": "تم التسليم", "cancelled": "ملغي",
    }
    return templates.TemplateResponse(request, "order_detail.html", {
        "admin_name": _name(request), "active": "orders",
        "order": to_obj(raw), "status_labels": status_labels,
        "all_statuses": list(status_labels.items()),
    })


@app.post("/orders/{order_id}/status")
async def update_order_status(
    order_id: int, request: Request,
    status: str = Form(...), note: str = Form(""), admin_notes: str = Form(""),
    estimated_time: str = Form(""), employee_name: str = Form(""),
):
    if not _logged(request):
        return _redirect_login()
    payload = {
        "status": status, "note": note or None,
        "admin_notes": admin_notes or None,
        "estimated_time": estimated_time or None,
        "employee_name": employee_name or _name(request),
    }
    _, err = await api_ex("put", f"/api/orders/{order_id}/status", token=_token(request), json=payload)
    if err:
        return RedirectResponse(f"/orders/{order_id}?error={_q(err)}", status_code=302)
    return RedirectResponse(f"/orders/{order_id}?success=تم+تحديث+حالة+الطلب+بنجاح", status_code=302)


# ── Maintenance ─────────────────────────────────────────────────────────────────

@app.get("/maintenance", response_class=HTMLResponse)
async def maintenance_list(request: Request, search: str = ""):
    if not _logged(request):
        return _redirect_login()
    params = {"limit": 300}
    if search:
        params["search"] = search
    raw = await api("get", "/api/maintenance/", token=_token(request), params=params) or []
    orders = to_obj(raw)
    maint_labels = {
        "received": "مستلم", "inspecting": "قيد الفحص", "repairing": "جاري الإصلاح",
        "waiting_part": "انتظار قطعة", "repaired": "تم الإصلاح",
        "ready": "جاهز للاستلام", "delivered": "تم التسليم",
    }
    maint_colors = {
        "received": "#6B7280", "inspecting": "#F59E0B", "repairing": "#8B5CF6",
        "waiting_part": "#EF4444", "repaired": "#06B6D4",
        "ready": "#10B981", "delivered": "#22C55E",
    }
    return templates.TemplateResponse(request, "maintenance.html", {
        "admin_name": _name(request), "active": "maintenance",
        "orders": orders, "maint_labels": maint_labels, "maint_colors": maint_colors,
        "search": search, "all_maint_statuses": list(maint_labels.items()),
    })


@app.post("/maintenance/add")
async def maintenance_add(
    request: Request,
    customer_name: str = Form(...), customer_phone: str = Form(...),
    device_type: str = Form(...),
    problem_description: str = Form(...), price: float = Form(0),
    notes: str = Form(""),
):
    if not _logged(request):
        return _redirect_login()
    payload = {
        "customer_name": customer_name, "customer_phone": customer_phone,
        "device_type": device_type, "problem_description": problem_description,
        "price": price, "notes": notes or None,
    }
    _, err = await api_ex("post", "/api/maintenance/", token=_token(request), json=payload)
    if err:
        return RedirectResponse(f"/maintenance?error={_q(err)}", status_code=302)
    return RedirectResponse("/maintenance?success=تم+إضافة+طلب+الصيانة+بنجاح", status_code=302)


@app.post("/maintenance/{order_id}/status")
async def update_maintenance_status(
    order_id: int, request: Request,
    maintenance_status: str = Form(...), note: str = Form(""),
    admin_notes: str = Form(""), estimated_time: str = Form(""),
):
    if not _logged(request):
        return _redirect_login()
    payload = {
        "maintenance_status": maintenance_status, "note": note or None,
        "admin_notes": admin_notes or None,
        "estimated_time": estimated_time or None,
        "employee_name": _name(request),
    }
    _, err = await api_ex("put", f"/api/maintenance/{order_id}/status", token=_token(request), json=payload)
    if err:
        return RedirectResponse(f"/maintenance?error={_q(err)}", status_code=302)
    return RedirectResponse("/maintenance?success=تم+تحديث+حالة+الصيانة+بنجاح", status_code=302)


# ── Reservations ────────────────────────────────────────────────────────────────

@app.get("/reservations", response_class=HTMLResponse)
async def reservations_list(request: Request):
    if not _logged(request):
        return _redirect_login()
    raw_res  = await api("get", "/api/reservations/", token=_token(request), params={"limit": 300}) or []
    raw_prod = await api("get", "/api/products/", token=_token(request), params={"status": "available", "limit": 200}) or []
    reservations       = to_obj(raw_res)
    available_products = to_obj(raw_prod)
    status_labels = {"pending":"قيد الانتظار","confirmed":"مؤكد","cancelled":"ملغي","completed":"مكتمل"}
    status_colors = {"pending":"#F59E0B","confirmed":"#22C55E","cancelled":"#EF4444","completed":"#06B6D4"}
    return templates.TemplateResponse(request, "reservations.html", {
        "admin_name": _name(request), "active": "reservations",
        "reservations": reservations, "status_labels": status_labels,
        "status_colors": status_colors, "available_products": available_products,
    })


@app.post("/reservations/add")
async def reservation_add(
    request: Request,
    customer_name: str = Form(...), customer_phone: str = Form(...),
    product_id: int = Form(...), notes: str = Form(""), days: int = Form(3),
):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("post", "/api/reservations/", token=_token(request),
                          json={"customer_name": customer_name, "customer_phone": customer_phone,
                                "product_id": product_id, "notes": notes or None, "days": days})
    if err:
        return RedirectResponse(f"/reservations?error={_q(err)}", status_code=302)
    return RedirectResponse("/reservations?success=تم+إضافة+الحجز+بنجاح", status_code=302)


@app.post("/reservations/{res_id}/cancel")
async def reservation_cancel(res_id: int, request: Request):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("put", f"/api/reservations/{res_id}/cancel", token=_token(request))
    if err:
        return RedirectResponse(f"/reservations?error={_q(err)}", status_code=302)
    return RedirectResponse("/reservations?success=تم+إلغاء+الحجز+بنجاح", status_code=302)


@app.post("/reservations/{res_id}/complete")
async def reservation_complete(res_id: int, request: Request):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("put", f"/api/reservations/{res_id}/complete", token=_token(request))
    if err:
        return RedirectResponse(f"/reservations?error={_q(err)}", status_code=302)
    return RedirectResponse("/reservations?success=تم+إكمال+الحجز+بنجاح", status_code=302)


# ── Customers ───────────────────────────────────────────────────────────────────

@app.get("/customers", response_class=HTMLResponse)
async def customers_list(request: Request, search: str = ""):
    if not _logged(request):
        return _redirect_login()
    params = {"limit": 300}
    if search:
        params["search"] = search
    raw = await api("get", "/api/customers/", token=_token(request), params=params) or []
    customers = to_obj(raw)
    return templates.TemplateResponse(request, "customers.html", {
        "admin_name": _name(request), "active": "customers",
        "customers": customers, "search": search,
    })


@app.post("/customers/add")
async def customer_add(
    request: Request,
    name: str = Form(...), phone: str = Form(...),
    password: str = Form(...),
):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("post", "/api/customers/", token=_token(request),
                          json={"name": name, "phone": phone or None,
                                "email": None, "password": password})
    if err:
        return RedirectResponse(f"/customers?error={_q(err)}", status_code=302)
    return RedirectResponse("/customers?success=تم+إضافة+العميل+بنجاح", status_code=302)


@app.post("/customers/{user_id}/toggle-active")
async def customer_toggle(user_id: int, request: Request):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("post", f"/api/customers/{user_id}/toggle-active", token=_token(request))
    if err:
        return RedirectResponse(f"/customers?error={_q(err)}", status_code=302)
    return RedirectResponse("/customers?success=تم+تغيير+حالة+العميل+بنجاح", status_code=302)


@app.post("/customers/{user_id}/verify")
async def customer_verify(user_id: int, request: Request):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("post", f"/api/customers/{user_id}/verify", token=_token(request))
    if err:
        return RedirectResponse(f"/customers?error={_q(err)}", status_code=302)
    return RedirectResponse("/customers?success=تم+توثيق+العميل+بنجاح", status_code=302)


@app.post("/customers/{user_id}/delete")
async def customer_delete(user_id: int, request: Request):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("delete", f"/api/customers/{user_id}", token=_token(request))
    if err:
        return RedirectResponse(f"/customers?error={_q(err)}", status_code=302)
    return RedirectResponse("/customers?success=تم+حذف+العميل+بنجاح", status_code=302)


# ── Inventory ───────────────────────────────────────────────────────────────────

@app.get("/inventory", response_class=HTMLResponse)
async def inventory_list(request: Request, search: str = "", status_filter: str = "", grade_filter: str = ""):
    if not _logged(request):
        return _redirect_login()
    params = {"limit": 300}
    if search:
        params["search"] = search
    if status_filter:
        params["status"] = status_filter
    if grade_filter:
        params["grade"] = grade_filter
    raw   = await api("get", "/api/inventory/", token=_token(request), params=params) or []
    stats = await api("get", "/api/inventory/stats/summary", token=_token(request)) or {}
    items = to_obj(raw)
    return templates.TemplateResponse(request, "inventory.html", {
        "admin_name": _name(request), "active": "inventory",
        "items": items, "search": search,
        "status_filter": status_filter, "grade_filter": grade_filter,
        "stats": stats,
    })


@app.post("/inventory/add")
async def inventory_add(
    request: Request,
    serial_number: str = Form(""), category: str = Form(...),
    brand: str = Form(""), model: str = Form(""), grade: str = Form("A"),
    price: float = Form(...), notes: str = Form(""),
):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("post", "/api/inventory/", token=_token(request),
                          json={"serial_number": serial_number or None, "category": category,
                                "brand": brand or None, "model": model or None,
                                "grade": grade, "price": price, "notes": notes or None})
    if err:
        return RedirectResponse(f"/inventory?error={_q(err)}", status_code=302)
    return RedirectResponse("/inventory?success=تم+إضافة+العنصر+للمخزون+بنجاح", status_code=302)


@app.post("/inventory/{item_id}/sell")
async def inventory_sell(item_id: int, request: Request):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("post", f"/api/inventory/{item_id}/sell", token=_token(request))
    if err:
        return RedirectResponse(f"/inventory?error={_q(err)}", status_code=302)
    return RedirectResponse("/inventory?success=تم+تسجيل+البيع+بنجاح", status_code=302)


@app.post("/inventory/{item_id}/return-stock")
async def inventory_return_stock(item_id: int, request: Request):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("post", f"/api/inventory/{item_id}/return-to-stock", token=_token(request))
    if err:
        return RedirectResponse(f"/inventory?error={_q(err)}", status_code=302)
    return RedirectResponse("/inventory?success=تم+إعادة+العنصر+للمخزون+بنجاح", status_code=302)


# ── Inspection ──────────────────────────────────────────────────────────────────

@app.get("/inspection", response_class=HTMLResponse)
async def inspection_list(request: Request, status: str = ""):
    if not _logged(request):
        return _redirect_login()
    params = {"limit": 200}
    if status:
        params["status"] = status
    raw = await api("get", "/api/inspection/", token=_token(request), params=params) or []
    return templates.TemplateResponse(request, "inspection.html", {
        "admin_name": _name(request), "active": "inspection",
        "requests": to_obj(raw),
    })


@app.post("/inspection/{req_id}/respond")
async def inspection_respond(
    req_id: int, request: Request,
    diagnosis: str = Form(...), estimated_price: str = Form(""),
    response_notes: str = Form(""),
    response_images: list[UploadFile] = File(default=[]),
):
    if not _logged(request):
        return _redirect_login()
    # Upload images via the backend API
    image_urls: list[str] = []
    for img in response_images:
        if img.filename:
            content = await img.read()
            import io as _io
            files_payload = {"file": (img.filename, _io.BytesIO(content), img.content_type or "image/jpeg")}
            upload_result, upload_err = await api_raw_upload(
                "/api/uploads/image", files_payload, token=_token(request)
            )
            if upload_result and isinstance(upload_result, dict):
                image_urls.append(upload_result.get("url", ""))
            elif upload_err:
                print(f"[inspection] image upload error: {upload_err}")
    _, err = await api_ex("post", f"/api/inspection/{req_id}/respond", token=_token(request),
              json={"staff_id": request.session.get("admin_id"),
                    "diagnosis": diagnosis,
                    "estimated_price": estimated_price or None,
                    "response_notes": response_notes or None,
                    "response_images": [u for u in image_urls if u]})
    if err:
        return RedirectResponse(f"/inspection?error={_q(err)}", status_code=302)
    return RedirectResponse("/inspection?success=تم+إرسال+رد+الفحص+بنجاح", status_code=302)


@app.post("/inspection/{req_id}/close")
async def inspection_close(req_id: int, request: Request):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("post", f"/api/inspection/{req_id}/close", token=_token(request))
    if err:
        return RedirectResponse(f"/inspection?error={_q(err)}", status_code=302)
    return RedirectResponse("/inspection?success=تم+إغلاق+طلب+الفحص+بنجاح", status_code=302)


# ── Referrals ───────────────────────────────────────────────────────────────────

@app.get("/referrals", response_class=HTMLResponse)
async def referrals_list(request: Request):
    if not _logged(request):
        return _redirect_login()
    raw = await api("get", "/api/referrals/all", token=_token(request), params={"limit": 300}) or []
    referrals = to_obj(raw)
    total    = len(referrals)
    verified = sum(1 for r in referrals if getattr(r, "is_verified", False))
    return templates.TemplateResponse(request, "referrals.html", {
        "admin_name": _name(request), "active": "referrals",
        "referrals": referrals, "total": total, "verified": verified,
        "top_referrers": [],
    })


# ── Warranty ────────────────────────────────────────────────────────────────────

@app.get("/warranty", response_class=HTMLResponse)
async def warranty_list(request: Request):
    if not _logged(request):
        return _redirect_login()
    raw = await api("get", "/api/warranty/", token=_token(request), params={"limit": 200}) or []
    warranties = to_obj(raw)
    return_requests = sum(
        1 for w in warranties
        if getattr(w, "is_return_requested", False) and not getattr(w, "return_resolved", False)
    )
    return templates.TemplateResponse(request, "warranty.html", {
        "admin_name": _name(request), "active": "warranty",
        "warranties": warranties, "return_requests": return_requests,
        "now_dt": datetime.utcnow(),
    })


@app.post("/warranty/add")
async def warranty_add(
    request: Request,
    product_name: str = Form(...), product_serial: str = Form(""),
    order_id: Optional[int] = Form(None), warranty_days: int = Form(7),
):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("post", "/api/warranty/", token=_token(request),
                          json={"product_name": product_name,
                                "product_serial": product_serial or None,
                                "order_id": order_id,
                                "warranty_days": warranty_days})
    if err:
        return RedirectResponse(f"/warranty?error={_q(err)}", status_code=302)
    return RedirectResponse("/warranty?success=تم+إضافة+الضمان+بنجاح", status_code=302)


@app.post("/warranty/{warranty_id}/resolve")
async def warranty_resolve(
    warranty_id: int, request: Request,
    approved: Optional[str] = Form(None),
    notes: str = Form(""),
):
    if not _logged(request):
        return _redirect_login()
    is_approved = approved == "true"
    _, err = await api_ex("post", f"/api/warranty/{warranty_id}/resolve-return", token=_token(request),
              json={"approved": is_approved, "notes": notes or "تم الحل من لوحة الإدارة"})
    if err:
        return RedirectResponse(f"/warranty?error={_q(err)}", status_code=302)
    return RedirectResponse("/warranty?success=تم+معالجة+طلب+الإرجاع+بنجاح", status_code=302)


@app.post("/warranty/{warranty_id}/edit")
async def warranty_edit(
    warranty_id: int, request: Request,
    product_name: str = Form(...),
    product_serial: str = Form(""),
    warranty_days: int = Form(7),
):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("put", f"/api/warranty/{warranty_id}/update", token=_token(request),
                          json={"product_name": product_name,
                                "product_serial": product_serial or None,
                                "warranty_days": warranty_days})
    if err:
        return RedirectResponse(f"/warranty?error={_q(err)}", status_code=302)
    return RedirectResponse("/warranty?success=تم+تحديث+الضمان+بنجاح", status_code=302)


@app.post("/warranty/{warranty_id}/delete")
async def warranty_delete(warranty_id: int, request: Request):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("delete", f"/api/warranty/{warranty_id}", token=_token(request))
    if err:
        return RedirectResponse(f"/warranty?error={_q(err)}", status_code=302)
    return RedirectResponse("/warranty?success=تم+حذف+الضمان+بنجاح", status_code=302)


# ── Staff ───────────────────────────────────────────────────────────────────────

@app.get("/staff", response_class=HTMLResponse)
async def staff_list(request: Request):
    if not _logged(request):
        return _redirect_login()
    raw = await api("get", "/api/staff/", token=_token(request)) or []
    return templates.TemplateResponse(request, "staff.html", {
        "admin_name": _name(request), "active": "staff",
        "staff": to_obj(raw),
    })


@app.post("/staff/add")
async def staff_add(
    request: Request,
    name: str = Form(...), phone: str = Form(""),
    role: str = Form("staff"), password: str = Form(...),
):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("post", "/api/staff/", token=_token(request),
                          json={"name": name, "phone": phone or None,
                                "email": None, "role": role, "password": password})
    if err:
        return RedirectResponse(f"/staff?error={_q(err)}", status_code=302)
    return RedirectResponse("/staff?success=تم+إضافة+الموظف+بنجاح", status_code=302)


@app.post("/staff/{user_id}/toggle-active")
async def staff_toggle(user_id: int, request: Request):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("post", f"/api/staff/{user_id}/toggle-active", token=_token(request))
    if err:
        return RedirectResponse(f"/staff?error={_q(err)}", status_code=302)
    return RedirectResponse("/staff?success=تم+تغيير+حالة+الموظف+بنجاح", status_code=302)


@app.post("/staff/{user_id}/edit")
async def staff_edit(
    user_id: int, request: Request,
    name: str = Form(...), phone: str = Form(""),
    role: str = Form("staff"),
    password: str = Form(""),
):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("put", f"/api/staff/{user_id}", token=_token(request),
                          json={"name": name, "phone": phone or None,
                                "email": None, "role": role,
                                "password": password or None})
    if err:
        return RedirectResponse(f"/staff?error={_q(err)}", status_code=302)
    return RedirectResponse("/staff?success=تم+تحديث+بيانات+الموظف+بنجاح", status_code=302)


@app.post("/staff/{user_id}/delete")
async def staff_delete(user_id: int, request: Request):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("delete", f"/api/staff/{user_id}", token=_token(request))
    if err:
        return RedirectResponse(f"/staff?error={_q(err)}", status_code=302)
    return RedirectResponse("/staff?success=تم+حذف+الموظف+بنجاح", status_code=302)


# ── Branches ────────────────────────────────────────────────────────────────────

@app.get("/branches", response_class=HTMLResponse)
async def branches_list(request: Request):
    if not _logged(request):
        return _redirect_login()
    raw_b = await api("get", "/api/branches/", token=_token(request)) or []
    raw_w = await api("get", "/api/branches/warehouses/all", token=_token(request)) or []
    return templates.TemplateResponse(request, "branches.html", {
        "admin_name": _name(request), "active": "branches",
        "branches": to_obj(raw_b), "warehouses": to_obj(raw_w),
    })


@app.post("/branches/add")
async def branch_add(
    request: Request,
    name: str = Form(...), city: str = Form(""),
    address: str = Form(""), phone: str = Form(""),
):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("post", "/api/branches/", token=_token(request),
                          json={"name": name, "city": city or None,
                                "address": address or None, "phone": phone or None})
    if err:
        return RedirectResponse(f"/branches?error={_q(err)}", status_code=302)
    return RedirectResponse("/branches?success=تم+إضافة+الفرع+بنجاح", status_code=302)


@app.post("/branches/{branch_id}/edit")
async def branch_edit(
    branch_id: int, request: Request,
    name: str = Form(...), city: str = Form(""),
    address: str = Form(""), phone: str = Form(""),
    is_active: Optional[str] = Form(None),
):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("put", f"/api/branches/{branch_id}", token=_token(request),
                          json={"name": name, "city": city or None,
                                "address": address or None, "phone": phone or None,
                                "is_active": bool(is_active)})
    if err:
        return RedirectResponse(f"/branches?error={_q(err)}", status_code=302)
    return RedirectResponse("/branches?success=تم+تحديث+الفرع+بنجاح", status_code=302)


@app.post("/branches/{branch_id}/delete")
async def branch_delete(branch_id: int, request: Request):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("delete", f"/api/branches/{branch_id}", token=_token(request))
    if err:
        return RedirectResponse(f"/branches?error={_q(err)}", status_code=302)
    return RedirectResponse("/branches?success=تم+حذف+الفرع+بنجاح", status_code=302)


# ── Notifications ───────────────────────────────────────────────────────────────

@app.get("/notifications", response_class=HTMLResponse)
async def notifications_list(request: Request):
    if not _logged(request):
        return _redirect_login()
    raw_notifs   = await api("get", "/api/notifications/all", token=_token(request)) or []
    raw_customers = await api("get", "/api/customers/", token=_token(request), params={"limit": 300}) or []
    notifs    = to_obj(raw_notifs)
    customers = to_obj(raw_customers)
    total  = len(notifs)
    unread = sum(1 for n in notifs if not getattr(n, "is_read", True))
    return templates.TemplateResponse(request, "notifications.html", {
        "admin_name": _name(request), "active": "notifications",
        "notifs": notifs, "total": total, "unread": unread,
        "customers": customers,
    })


@app.post("/notifications/send")
async def notification_send(
    request: Request,
    title: str = Form(...), message: str = Form(...),
    user_id: Optional[int] = Form(None), broadcast: Optional[str] = Form(None),
):
    if not _logged(request):
        return _redirect_login()
    if broadcast:
        _, err = await api_ex("post", "/api/notifications/broadcast", token=_token(request),
                  params={"title": title, "body": message})
    elif user_id:
        _, err = await api_ex("post", "/api/notifications/send", token=_token(request),
                  json={"user_id": user_id, "title": title, "body": message})
    else:
        return RedirectResponse("/notifications?error=يجب+اختيار+عميل+أو+تفعيل+الإرسال+الجماعي", status_code=302)
    if err:
        return RedirectResponse(f"/notifications?error={_q(err)}", status_code=302)
    return RedirectResponse("/notifications?success=تم+إرسال+الإشعار+بنجاح", status_code=302)


@app.post("/notifications/{notif_id}/delete")
async def notification_delete(notif_id: int, request: Request):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("delete", f"/api/notifications/{notif_id}", token=_token(request))
    if err:
        return RedirectResponse(f"/notifications?error={_q(err)}", status_code=302)
    return RedirectResponse("/notifications?success=تم+حذف+الإشعار+بنجاح", status_code=302)


# ── Reports ─────────────────────────────────────────────────────────────────────

@app.get("/reports", response_class=HTMLResponse)
async def reports_page(request: Request, period: str = "month"):
    if not _logged(request):
        return _redirect_login()
    params = {"period": period}
    sales  = to_obj(await api("get", "/api/reports/sales",       token=_token(request), params=params) or {})
    profit = to_obj(await api("get", "/api/reports/profit",      token=_token(request), params=params) or {})
    inv    = to_obj(await api("get", "/api/reports/inventory",   token=_token(request)) or {})
    maint  = to_obj(await api("get", "/api/reports/maintenance", token=_token(request), params=params) or {})
    ref    = to_obj(await api("get", "/api/reports/referrals",   token=_token(request)) or {})
    war    = to_obj(await api("get", "/api/reports/warranty",    token=_token(request)) or {})
    return templates.TemplateResponse(request, "reports.html", {
        "admin_name": _name(request), "active": "reports",
        "period": period,
        "sales": sales, "profit": profit, "inv": inv,
        "maint": maint, "ref": ref, "war": war,
    })


# ── Wallet ──────────────────────────────────────────────────────────────────────

@app.get("/wallet", response_class=HTMLResponse)
async def wallet_page(request: Request, search: str = ""):
    if not _logged(request):
        return _redirect_login()
    params: dict = {"limit": 300}
    if search:
        params["search"] = search
    raw_customers = await api("get", "/api/customers/", token=_token(request), params=params) or []
    users = to_obj(raw_customers)
    if search and users:
        q = search.lower()
        users = [u for u in users if q in (getattr(u, "name", "") or "").lower()
                 or q in (getattr(u, "phone", "") or "").lower()]
    return templates.TemplateResponse(request, "wallet.html", {
        "admin_name": _name(request), "active": "wallet",
        "users": users,
        "search": search,
    })


@app.post("/wallet/{user_id}/credit")
async def wallet_credit(
    user_id: int, request: Request,
    amount: float = Form(...), note: str = Form(""),
):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("post", "/api/wallet/credit", token=_token(request),
              json={"user_id": user_id, "amount": amount,
                    "transaction_type": "credit",
                    "reason": note or "إضافة رصيد من لوحة التحكم"})
    if err:
        return RedirectResponse(f"/wallet?error={_q(err)}", status_code=302)
    return RedirectResponse("/wallet?success=تم+إضافة+الرصيد+بنجاح", status_code=302)


@app.post("/wallet/{user_id}/debit")
async def wallet_debit(
    user_id: int, request: Request,
    amount: float = Form(...), note: str = Form(""),
):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("post", "/api/wallet/debit", token=_token(request),
              json={"user_id": user_id, "amount": amount,
                    "transaction_type": "debit",
                    "reason": note or "خصم رصيد من لوحة التحكم"})
    if err:
        return RedirectResponse(f"/wallet?error={_q(err)}", status_code=302)
    return RedirectResponse("/wallet?success=تم+خصم+الرصيد+بنجاح", status_code=302)


# ── Admin Profile ───────────────────────────────────────────────────────────────

@app.get("/profile", response_class=HTMLResponse)
async def profile_page(request: Request):
    if not _logged(request):
        return _redirect_login()
    raw = await api("get", "/api/auth/me", token=_token(request)) or {}
    admin = to_obj(raw) if raw else None
    return templates.TemplateResponse(request, "profile.html", {
        "admin_name": _name(request),
        "active": "profile",
        "admin": admin,
        "success": request.query_params.get("success"),
        "error": request.query_params.get("error"),
    })


@app.post("/profile", response_class=HTMLResponse)
async def profile_update(
    request: Request,
    name: str = Form(...),
    phone: str = Form(""),
    current_password: str = Form(""),
    new_password: str = Form(""),
    confirm_password: str = Form(""),
):
    if not _logged(request):
        return _redirect_login()

    if new_password and new_password != confirm_password:
        return RedirectResponse("/profile?error=كلمتا+المرور+غير+متطابقتين", status_code=302)

    payload: dict = {"name": name}
    if phone.strip():
        payload["phone"] = phone.strip()
    else:
        payload["phone"] = None
    if new_password:
        if not current_password:
            return RedirectResponse("/profile?error=كلمة+المرور+الحالية+مطلوبة", status_code=302)
        payload["current_password"] = current_password
        payload["new_password"] = new_password

    result = await api("put", "/api/auth/profile", token=_token(request), json=payload)
    if result:
        request.session["admin_name"] = result.get("name", _name(request))
        return RedirectResponse("/profile?success=1", status_code=302)
    return RedirectResponse("/profile?error=فشل+التحديث.+تحقق+من+البيانات", status_code=302)


# ── Audit Log ───────────────────────────────────────────────────────────────────

@app.get("/audit", response_class=HTMLResponse)
async def audit_page(request: Request):
    if not _logged(request):
        return _redirect_login()
    raw = await api("get", "/api/audit/", token=_token(request), params={"limit": 200}) or []
    return templates.TemplateResponse(request, "audit.html", {
        "admin_name": _name(request), "active": "audit",
        "logs": to_obj(raw),
    })


@app.get("/audit/feed")
async def audit_feed(request: Request):
    from fastapi.responses import JSONResponse
    if not _logged(request):
        return JSONResponse({"logs": []})
    raw = await api("get", "/api/audit/", token=_token(request), params={"limit": 50}) or []
    logs_out = []
    for log in (raw if isinstance(raw, list) else []):
        if isinstance(log, dict):
            logs_out.append({
                "id": log.get("id"),
                "user_name": (log.get("user") or {}).get("name", "نظام") if isinstance(log.get("user"), dict) else "نظام",
                "user_role": log.get("user_role"),
                "action": log.get("action"),
                "entity_type": log.get("entity_type"),
                "entity_id": log.get("entity_id"),
                "description": log.get("description"),
                "ip_address": log.get("ip_address"),
                "created_at": log.get("created_at"),
            })
    return JSONResponse({"logs": logs_out})


# ── Announcements ────────────────────────────────────────────────────────────

@app.get("/announcements", response_class=HTMLResponse)
async def announcements_list(request: Request):
    if not _logged(request):
        return _redirect_login()
    raw = await api("get", "/api/announcements/", token=_token(request),
                    params={"active_only": "false"}) or []
    return templates.TemplateResponse(request, "announcements.html", {
        "admin_name": _name(request), "active": "announcements",
        "announcements": to_obj(raw),
    })


@app.post("/announcements/add")
async def announcement_add(
    request: Request,
    title: str = Form(...), body: str = Form(...),
    announcement_type: str = Form("info"),
    image_url: str = Form(""), action_url: str = Form(""),
    is_pinned: Optional[str] = Form(None),
    is_active: Optional[str] = Form(None),
):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("post", "/api/announcements/", token=_token(request),
                          json={
                              "title": title, "body": body,
                              "announcement_type": announcement_type,
                              "image_url": image_url or None,
                              "action_url": action_url or None,
                              "is_pinned": bool(is_pinned),
                              "is_active": bool(is_active),
                          })
    if err:
        return RedirectResponse(f"/announcements?error={_q(err)}", status_code=302)
    return RedirectResponse("/announcements?success=تم+إضافة+الإعلان+بنجاح", status_code=302)


@app.post("/announcements/{ann_id}/toggle")
async def announcement_toggle(ann_id: int, request: Request):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("post", f"/api/announcements/{ann_id}/toggle", token=_token(request))
    if err:
        return RedirectResponse(f"/announcements?error={_q(err)}", status_code=302)
    return RedirectResponse("/announcements?success=تم+تغيير+حالة+الإعلان+بنجاح", status_code=302)


@app.post("/announcements/{ann_id}/delete")
async def announcement_delete(ann_id: int, request: Request):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("delete", f"/api/announcements/{ann_id}", token=_token(request))
    if err:
        return RedirectResponse(f"/announcements?error={_q(err)}", status_code=302)
    return RedirectResponse("/announcements?success=تم+حذف+الإعلان+بنجاح", status_code=302)


# ── Export to Excel ───────────────────────────────────────────────────────────
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from fastapi.responses import StreamingResponse
from datetime import date as _date


def _make_wb(title: str):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    ws.sheet_view.rightToLeft = True
    return wb, ws


def _style_header(ws, headers: list):
    fill = PatternFill("solid", fgColor="1A73E8")
    font = Font(bold=True, color="FFFFFF", size=12, name="Cairo")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22


def _set_widths(ws, widths: list):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _wb_response(wb, filename: str):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/export/customers")
async def export_customers(request: Request):
    if not _logged(request):
        return _redirect_login()
    raw = await api("get", "/api/customers/", token=_token(request), params={"limit": 5000}) or []
    wb, ws = _make_wb("العملاء")
    _style_header(ws, ["#", "الاسم", "رقم الجوال", "البريد الإلكتروني", "رصيد المحفظة", "الحالة", "التوثيق", "تاريخ التسجيل"])
    _set_widths(ws, [6, 25, 18, 30, 15, 12, 12, 20])
    for i, c in enumerate(raw, 1):
        created = (c.get("created_at") or "")[:16].replace("T", " ")
        ws.append([i, c.get("name",""), c.get("phone",""), c.get("email",""),
                   c.get("wallet_balance", 0),
                   "نشط" if c.get("is_active") else "موقوف",
                   "موثّق" if c.get("is_verified") else "غير موثّق", created])
    return _wb_response(wb, f"customers_{_date.today()}.xlsx")


@app.get("/export/staff")
async def export_staff_excel(request: Request):
    if not _logged(request):
        return _redirect_login()
    raw = await api("get", "/api/staff/", token=_token(request)) or []
    wb, ws = _make_wb("الموظفون")
    _style_header(ws, ["#", "الاسم", "رقم الجوال", "البريد الإلكتروني", "الدور", "الحالة", "تاريخ الإضافة"])
    _set_widths(ws, [6, 25, 18, 30, 15, 12, 20])
    roles = {"staff": "موظف", "branch_manager": "مدير فرع", "admin": "مدير عام"}
    for i, s in enumerate(raw, 1):
        created = (s.get("created_at") or "")[:16].replace("T", " ")
        ws.append([i, s.get("name",""), s.get("phone",""), s.get("email",""),
                   roles.get(s.get("role",""), s.get("role","")),
                   "نشط" if s.get("is_active") else "معطّل", created])
    return _wb_response(wb, f"staff_{_date.today()}.xlsx")


@app.get("/export/products")
async def export_products_excel(request: Request):
    if not _logged(request):
        return _redirect_login()
    raw = await api("get", "/api/products/", token=_token(request), params={"limit": 5000}) or []
    wb, ws = _make_wb("المنتجات")
    _style_header(ws, ["#", "الاسم", "الاسم بالعربي", "الماركة", "الموديل", "الفئة", "السعر", "الكمية", "الحالة", "تاريخ الإضافة"])
    _set_widths(ws, [6, 25, 25, 15, 18, 15, 12, 10, 12, 20])
    status_labels = {"available":"متوفر","reserved":"محجوز","sold":"مباع","unavailable":"غير متوفر"}
    for i, p in enumerate(raw, 1):
        created = (p.get("created_at") or "")[:16].replace("T", " ")
        ws.append([i, p.get("name",""), p.get("name_ar",""), p.get("brand",""), p.get("model",""),
                   p.get("category",""), p.get("price",0), p.get("quantity",0),
                   status_labels.get(p.get("status",""), p.get("status","")), created])
    return _wb_response(wb, f"products_{_date.today()}.xlsx")


@app.get("/export/orders")
async def export_orders_excel(request: Request):
    if not _logged(request):
        return _redirect_login()
    raw = await api("get", "/api/orders/", token=_token(request), params={"limit": 5000}) or []
    wb, ws = _make_wb("الطلبات")
    _style_header(ws, ["#", "رقم الطلب", "اسم العميل", "الجوال", "المنتج", "السعر", "الحالة", "ملاحظات", "تاريخ الطلب"])
    _set_widths(ws, [6, 10, 25, 18, 25, 12, 15, 30, 20])
    status_labels = {"received":"مستلم","reviewing":"قيد المراجعة","confirmed":"مؤكد",
                     "preparing":"جاري التحضير","shipped":"تم الشحن","on_the_way":"في الطريق",
                     "delivered":"تم التسليم","cancelled":"ملغي"}
    for i, o in enumerate(raw, 1):
        created = (o.get("created_at") or "")[:16].replace("T", " ")
        product = o.get("product") or {}
        user = o.get("user") or {}
        ws.append([i, o.get("id",""),
                   o.get("customer_name") or user.get("name",""),
                   o.get("customer_phone") or user.get("phone",""),
                   product.get("name","") if isinstance(product, dict) else "",
                   o.get("total_price", o.get("price", 0)),
                   status_labels.get(o.get("status",""), o.get("status","")),
                   o.get("admin_notes",""), created])
    return _wb_response(wb, f"orders_{_date.today()}.xlsx")


@app.get("/export/maintenance")
async def export_maintenance_excel(request: Request):
    if not _logged(request):
        return _redirect_login()
    raw = await api("get", "/api/maintenance/", token=_token(request), params={"limit": 5000}) or []
    wb, ws = _make_wb("الصيانة")
    _style_header(ws, ["#", "رقم الطلب", "اسم العميل", "الجوال", "نوع الجهاز", "المشكلة", "السعر", "الحالة", "تاريخ الاستلام"])
    _set_widths(ws, [6, 10, 25, 18, 20, 35, 12, 18, 20])
    maint_labels = {"received":"مستلم","inspecting":"قيد الفحص","repairing":"جاري الإصلاح",
                    "waiting_part":"انتظار قطعة","repaired":"تم الإصلاح",
                    "ready":"جاهز للاستلام","delivered":"تم التسليم"}
    for i, o in enumerate(raw, 1):
        created = (o.get("created_at") or "")[:16].replace("T", " ")
        status = o.get("maintenance_status", o.get("status",""))
        ws.append([i, o.get("id",""), o.get("customer_name",""), o.get("customer_phone",""),
                   o.get("device_type",""), o.get("problem_description",""),
                   o.get("price",0), maint_labels.get(status, status), created])
    return _wb_response(wb, f"maintenance_{_date.today()}.xlsx")


# ══════════════════════════════════════════════════════════════════════════════
#  LOYALTY POINTS
# ══════════════════════════════════════════════════════════════════════════════
@app.get("/loyalty")
async def loyalty_list(request: Request):
    if not _logged(request):
        return _redirect_login()
    accounts = to_obj(await api("get", "/api/loyalty/all", token=_token(request)) or [])
    return templates.TemplateResponse(request, "loyalty.html", {
        "admin_name": _name(request),
        "active": "loyalty",
        "accounts": accounts,
        "show_add_form": False,
        "success": request.query_params.get("success"),
        "error": request.query_params.get("error"),
    })


@app.get("/loyalty/add")
async def loyalty_add_form(request: Request):
    if not _logged(request):
        return _redirect_login()
    accounts = to_obj(await api("get", "/api/loyalty/all", token=_token(request)) or [])
    return templates.TemplateResponse(request, "loyalty.html", {
        "admin_name": _name(request),
        "active": "loyalty",
        "accounts": accounts,
        "show_add_form": True,
        "success": None,
        "error": None,
    })


@app.post("/loyalty/add")
async def loyalty_add_points(request: Request, phone: str = Form(...), points: int = Form(1), reason: str = Form("بيع شاشة")):
    if not _logged(request):
        return _redirect_login()
    # Find user by phone
    users = await api("get", "/api/customers/", token=_token(request), params={"phone": phone, "limit": 5}) or []
    if not users:
        return RedirectResponse(f"/loyalty?error={_q('لم يتم العثور على عميل بهذا الرقم')}", status_code=302)
    user_id = users[0].get("id") if isinstance(users[0], dict) else getattr(users[0], "id", None)
    result = await api("post", "/api/loyalty/add-points", token=_token(request), json={
        "user_id": user_id, "points": points, "reason": reason
    })
    msg = result.get("message", "تم") if isinstance(result, dict) else "تم"
    return RedirectResponse(f"/loyalty?success={_q(msg)}", status_code=302)


@app.post("/loyalty/reset/{user_id}")
async def loyalty_reset(request: Request, user_id: int, reason: str = Form("تسليم شاشة مجانية")):
    if not _logged(request):
        return _redirect_login()
    result, err = await api_ex("post", "/api/loyalty/reset", token=_token(request), json={"user_id": user_id, "reason": reason})
    if err:
        return RedirectResponse(f"/loyalty?error={_q(err)}", status_code=302)
    msg = result.get("message", "تم التصفير") if isinstance(result, dict) else "تم التصفير"
    return RedirectResponse(f"/loyalty?success={_q(msg)}", status_code=302)


@app.get("/loyalty/user/{user_id}")
async def loyalty_user_transactions(request: Request, user_id: int):
    if not _logged(request):
        return _redirect_login()
    txs = to_obj(await api("get", f"/api/loyalty/transactions/{user_id}", token=_token(request)) or [])
    acc = to_obj(await api("get", f"/api/loyalty/account/{user_id}", token=_token(request)) or {})
    return templates.TemplateResponse(request, "loyalty_user.html", {
        "admin_name": _name(request),
        "active": "loyalty",
        "transactions": txs,
        "account": acc,
        "user_id": user_id,
    })


# ══════════════════════════════════════════════════════════════════════════════
#  SHORTAGE REQUESTS
# ══════════════════════════════════════════════════════════════════════════════
@app.get("/shortage-requests")
async def shortage_list(request: Request, status: str = "", view: str = "list"):
    if not _logged(request):
        return _redirect_login()
    params = {"limit": 500}
    if status:
        params["status"] = status
    requests_data = to_obj(await api("get", "/api/shortage-requests/", token=_token(request), params=params) or [])
    groups = to_obj(await api("get", "/api/shortage-requests/grouped", token=_token(request)) or [])
    return templates.TemplateResponse(request, "shortage_requests.html", {
        "admin_name": _name(request),
        "active": "shortage-requests",
        "requests": requests_data,
        "groups": groups,
        "status_filter": status,
        "view": view,
        "success": request.query_params.get("success"),
        "error": request.query_params.get("error"),
    })


@app.post("/shortage-requests/{req_id}/notify")
async def shortage_notify(request: Request, req_id: int):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("put", f"/api/shortage-requests/{req_id}/notify", token=_token(request))
    if err:
        return RedirectResponse(f"/shortage-requests?error={_q(err)}", status_code=302)
    return RedirectResponse(f"/shortage-requests?success={_q('تم تحديث الحالة وإشعار العميل')}", status_code=302)


@app.post("/shortage-requests/{req_id}/purchased")
async def shortage_purchased(request: Request, req_id: int):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("put", f"/api/shortage-requests/{req_id}/purchased", token=_token(request))
    if err:
        return RedirectResponse(f"/shortage-requests?error={_q(err)}", status_code=302)
    return RedirectResponse(f"/shortage-requests?success={_q('تم تسجيل الشراء بنجاح')}", status_code=302)


@app.post("/shortage-requests/{req_id}/close")
async def shortage_close(request: Request, req_id: int):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("put", f"/api/shortage-requests/{req_id}/close", token=_token(request))
    if err:
        return RedirectResponse(f"/shortage-requests?error={_q(err)}", status_code=302)
    return RedirectResponse(f"/shortage-requests?success={_q('تم إغلاق الطلب بنجاح')}", status_code=302)


@app.post("/shortage-requests/batch-notify")
async def shortage_batch_notify(request: Request):
    if not _logged(request):
        return _redirect_login()
    form = await request.form()
    ids = [int(x) for x in form.getlist("ids") if x]
    _, err = await api_ex("put", "/api/shortage-requests/batch-notify", token=_token(request), json={"request_ids": ids})
    if err:
        return RedirectResponse(f"/shortage-requests?error={_q(err)}", status_code=302)
    return RedirectResponse(f"/shortage-requests?success={_q(f'تم إشعار {len(ids)} طلب بنجاح')}", status_code=302)


# ══════════════════════════════════════════════════════════════════════════════
#  AUCTIONS
# ══════════════════════════════════════════════════════════════════════════════
@app.get("/auctions")
async def auctions_list(request: Request, status: str = ""):
    if not _logged(request):
        return _redirect_login()
    params = {"limit": 200}
    if status:
        params["status"] = status
    auctions = to_obj(await api("get", "/api/auctions/admin/all", token=_token(request), params=params) or [])
    return templates.TemplateResponse(request, "auctions.html", {
        "admin_name": _name(request),
        "active": "auctions",
        "auctions": auctions,
        "status_filter": status,
        "success": request.query_params.get("success"),
        "error": request.query_params.get("error"),
    })


@app.get("/auctions/{auction_id}")
async def auction_detail(request: Request, auction_id: int):
    if not _logged(request):
        return _redirect_login()
    auction = to_obj(await api("get", f"/api/auctions/{auction_id}", token=_token(request)) or {})
    return templates.TemplateResponse(request, "auction_detail.html", {
        "admin_name": _name(request),
        "active": "auctions",
        "auction": auction,
    })


@app.post("/auctions/{auction_id}/activate")
async def auction_activate(request: Request, auction_id: int):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("put", f"/api/auctions/{auction_id}/activate", token=_token(request), params={"days": 7})
    if err:
        return RedirectResponse(f"/auctions?error={_q(err)}", status_code=302)
    return RedirectResponse(f"/auctions?success={_q('تم تفعيل المزاد بنجاح')}", status_code=302)


@app.post("/auctions/{auction_id}/reject")
async def auction_reject(request: Request, auction_id: int):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("delete", f"/api/auctions/{auction_id}", token=_token(request))
    if err:
        return RedirectResponse(f"/auctions?error={_q(err)}", status_code=302)
    return RedirectResponse(f"/auctions?success={_q('تم رفض المزاد بنجاح')}", status_code=302)


@app.post("/auctions/{auction_id}/close")
async def auction_close(request: Request, auction_id: int):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("put", f"/api/auctions/{auction_id}/close", token=_token(request))
    if err:
        return RedirectResponse(f"/auctions?error={_q(err)}", status_code=302)
    return RedirectResponse(f"/auctions?success={_q('تم إغلاق المزاد بنجاح')}", status_code=302)


# ══════════════════════════════════════════════════════════════════════════════
#  SECRET DEALS
# ══════════════════════════════════════════════════════════════════════════════
@app.get("/secret-deals")
async def secret_deals_list(request: Request, status: str = ""):
    if not _logged(request):
        return _redirect_login()
    params = {"limit": 100}
    if status:
        params["status"] = status
    deals = to_obj(await api("get", "/api/secret-deals/", token=_token(request), params=params) or [])
    return templates.TemplateResponse(request, "secret_deals.html", {
        "admin_name": _name(request),
        "active": "secret-deals",
        "deals": deals,
        "status_filter": status,
        "success": request.query_params.get("success"),
        "error": request.query_params.get("error"),
    })


@app.get("/secret-deals/add")
async def secret_deal_add_form(request: Request):
    if not _logged(request):
        return _redirect_login()
    return templates.TemplateResponse(request, "secret_deal_add.html", {
        "admin_name": _name(request),
        "active": "secret-deals",
    })


@app.post("/secret-deals/add")
async def secret_deal_create(
    request: Request,
    title: str = Form(...),
    supplier_name: Optional[str] = Form(None),
    supplier_phone: Optional[str] = Form(None),
    total_quantity: Optional[int] = Form(0),
    price_per_unit: Optional[float] = Form(None),
    total_price: Optional[float] = Form(None),
    admin_notes: Optional[str] = Form(None),
):
    if not _logged(request):
        return _redirect_login()
    result, err = await api_ex("post", "/api/secret-deals/", token=_token(request), json={
        "title": title, "supplier_name": supplier_name, "supplier_phone": supplier_phone,
        "total_quantity": total_quantity, "price_per_unit": price_per_unit,
        "total_price": total_price, "admin_notes": admin_notes,
    })
    if err:
        return RedirectResponse(f"/secret-deals?error={_q(err)}", status_code=302)
    deal_id = result.get("id") if isinstance(result, dict) else None
    return RedirectResponse(f"/secret-deals/{deal_id}" if deal_id else f"/secret-deals?success={_q('تم إنشاء الصفقة بنجاح')}", status_code=302)


@app.get("/secret-deals/{deal_id}")
async def secret_deal_detail(request: Request, deal_id: int):
    if not _logged(request):
        return _redirect_login()
    deal = to_obj(await api("get", f"/api/secret-deals/{deal_id}", token=_token(request)) or {})
    return templates.TemplateResponse(request, "secret_deal_detail.html", {
        "admin_name": _name(request),
        "active": "secret-deals",
        "deal": deal,
        "success": request.query_params.get("success"),
        "error": request.query_params.get("error"),
    })


@app.get("/secret-deals/{deal_id}/upload")
async def secret_deal_upload_form(request: Request, deal_id: int):
    if not _logged(request):
        return _redirect_login()
    deal = to_obj(await api("get", f"/api/secret-deals/{deal_id}", token=_token(request)) or {})
    return templates.TemplateResponse(request, "secret_deal_upload.html", {
        "admin_name": _name(request),
        "active": "secret-deals",
        "deal": deal,
    })


@app.post("/secret-deals/{deal_id}/upload")
async def secret_deal_upload(request: Request, deal_id: int, image_urls: str = Form("")):
    if not _logged(request):
        return _redirect_login()
    urls = [u.strip() for u in image_urls.split("\n") if u.strip()]
    result, err = await api_ex("post", f"/api/secret-deals/{deal_id}/images", token=_token(request), json={"image_urls": urls})
    if err:
        return RedirectResponse(f"/secret-deals/{deal_id}?error={_q(err)}", status_code=302)
    msg = result.get("message", "تمت إضافة الصور بنجاح") if isinstance(result, dict) else "تمت إضافة الصور بنجاح"
    return RedirectResponse(f"/secret-deals/{deal_id}?success={_q(msg)}", status_code=302)


@app.post("/secret-deals/{deal_id}/status")
async def secret_deal_update_status(request: Request, deal_id: int, status: str = Form(...)):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("put", f"/api/secret-deals/{deal_id}/status", token=_token(request), params={"status": status})
    if err:
        return RedirectResponse(f"/secret-deals?error={_q(err)}", status_code=302)
    return RedirectResponse(f"/secret-deals?success={_q('تم تحديث حالة الصفقة بنجاح')}", status_code=302)


# ══════════════════════════════════════════════════════════════════════════════
#  ENGINEERING SUPPORT
# ══════════════════════════════════════════════════════════════════════════════
@app.get("/eng-support")
async def eng_support_list(request: Request, status: str = ""):
    if not _logged(request):
        return _redirect_login()
    params = {"limit": 100}
    if status:
        params["status"] = status
    posts = to_obj(await api("get", "/api/eng-support/", token=_token(request), params=params) or [])
    return templates.TemplateResponse(request, "eng_support.html", {
        "admin_name": _name(request),
        "active": "eng-support",
        "posts": posts,
        "status_filter": status,
        "success": request.query_params.get("success"),
        "error": request.query_params.get("error"),
    })


@app.get("/eng-support/{post_id}")
async def eng_support_detail(request: Request, post_id: int):
    if not _logged(request):
        return _redirect_login()
    post = to_obj(await api("get", f"/api/eng-support/{post_id}", token=_token(request)) or {})
    return templates.TemplateResponse(request, "eng_support_detail.html", {
        "admin_name": _name(request),
        "active": "eng-support",
        "post": post,
        "success": request.query_params.get("success"),
        "error": request.query_params.get("error"),
    })


@app.post("/eng-support/{post_id}/pin")
async def eng_support_pin(request: Request, post_id: int, is_pinned: bool = Form(True)):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("put", f"/api/eng-support/{post_id}", token=_token(request), json={"is_pinned": is_pinned})
    if err:
        return RedirectResponse(f"/eng-support?error={_q(err)}", status_code=302)
    return RedirectResponse(f"/eng-support?success={_q('تم تثبيت المنشور بنجاح')}", status_code=302)


@app.post("/eng-support/{post_id}/delete")
async def eng_support_delete(request: Request, post_id: int):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("delete", f"/api/eng-support/{post_id}", token=_token(request))
    if err:
        return RedirectResponse(f"/eng-support?error={_q(err)}", status_code=302)
    return RedirectResponse(f"/eng-support?success={_q('تم حذف المنشور بنجاح')}", status_code=302)


# ══════════════════════════════════════════════════════════════════════════════
#  COMPLAINTS
# ══════════════════════════════════════════════════════════════════════════════
@app.get("/complaints")
async def complaints_list(request: Request, status: str = "", type: str = ""):
    if not _logged(request):
        return _redirect_login()
    params = {"limit": 200}
    if status:
        params["status"] = status
    if type:
        params["complaint_type"] = type
    complaints = to_obj(await api("get", "/api/complaints/", token=_token(request), params=params) or [])
    unread = await api("get", "/api/complaints/unread-count", token=_token(request)) or {}
    unread_count = unread.get("unread_count", 0) if isinstance(unread, dict) else 0
    return templates.TemplateResponse(request, "complaints.html", {
        "admin_name": _name(request),
        "active": "complaints",
        "complaints": complaints,
        "status_filter": status,
        "type_filter": type,
        "unread_count": unread_count,
        "success": request.query_params.get("success"),
        "error": request.query_params.get("error"),
    })


@app.get("/complaints/{complaint_id}")
async def complaint_detail(request: Request, complaint_id: int):
    if not _logged(request):
        return _redirect_login()
    complaint = to_obj(await api("get", f"/api/complaints/{complaint_id}", token=_token(request)) or {})
    return templates.TemplateResponse(request, "complaint_detail.html", {
        "admin_name": _name(request),
        "active": "complaints",
        "complaint": complaint,
        "success": request.query_params.get("success"),
        "error": request.query_params.get("error"),
    })


@app.post("/complaints/{complaint_id}/reply")
async def complaint_reply(request: Request, complaint_id: int, reply: str = Form(...)):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("put", f"/api/complaints/{complaint_id}/reply", token=_token(request), json={"reply": reply})
    if err:
        return RedirectResponse(f"/complaints/{complaint_id}?error={_q(err)}", status_code=302)
    return RedirectResponse(f"/complaints/{complaint_id}?success={_q('تم إرسال الرد بنجاح')}", status_code=302)


@app.post("/complaints/{complaint_id}/resolve")
async def complaint_resolve(request: Request, complaint_id: int):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("put", f"/api/complaints/{complaint_id}/resolve", token=_token(request))
    if err:
        return RedirectResponse(f"/complaints?error={_q(err)}", status_code=302)
    return RedirectResponse(f"/complaints?success={_q('تم حل الشكوى بنجاح')}", status_code=302)


@app.post("/complaints/{complaint_id}/archive")
async def complaint_archive(request: Request, complaint_id: int):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("put", f"/api/complaints/{complaint_id}/archive", token=_token(request))
    if err:
        return RedirectResponse(f"/complaints?error={_q(err)}", status_code=302)
    return RedirectResponse(f"/complaints?success={_q('تمت أرشفة الشكوى بنجاح')}", status_code=302)


# ══════════════════════════════════════════════════════════════════════════════
#  PURCHASE INVOICES
# ══════════════════════════════════════════════════════════════════════════════
@app.get("/purchase-invoices")
async def purchase_invoices_list(request: Request):
    if not _logged(request):
        return _redirect_login()
    invoices = to_obj(await api("get", "/api/purchase-invoices/", token=_token(request), params={"limit": 300}) or [])
    summary = await api("get", "/api/purchase-invoices/summary", token=_token(request)) or {}
    total_amount = summary.get("total_purchased", 0) if isinstance(summary, dict) else 0
    total_drawer = summary.get("total_from_drawer", 0) if isinstance(summary, dict) else 0
    total_owner = summary.get("total_from_owner", 0) if isinstance(summary, dict) else 0
    return templates.TemplateResponse(request, "purchase_invoices.html", {
        "admin_name": _name(request),
        "active": "purchase-invoices",
        "invoices": invoices,
        "total_amount": total_amount,
        "total_drawer": total_drawer,
        "total_owner": total_owner,
        "show_add_form": False,
        "success": request.query_params.get("success"),
        "error": request.query_params.get("error"),
    })


@app.get("/purchase-invoices/add")
async def purchase_invoice_add_form(request: Request):
    if not _logged(request):
        return _redirect_login()
    invoices = to_obj(await api("get", "/api/purchase-invoices/", token=_token(request), params={"limit": 300}) or [])
    summary = await api("get", "/api/purchase-invoices/summary", token=_token(request)) or {}
    return templates.TemplateResponse(request, "purchase_invoices.html", {
        "admin_name": _name(request),
        "active": "purchase-invoices",
        "invoices": invoices,
        "total_amount": summary.get("total_purchased", 0) if isinstance(summary, dict) else 0,
        "total_drawer": summary.get("total_from_drawer", 0) if isinstance(summary, dict) else 0,
        "total_owner": summary.get("total_from_owner", 0) if isinstance(summary, dict) else 0,
        "show_add_form": True,
        "success": None,
        "error": None,
    })


@app.post("/purchase-invoices/add")
async def purchase_invoice_create(request: Request):
    if not _logged(request):
        return _redirect_login()
    form = await request.form()
    supplier_name = form.get("supplier_name", "")
    supplier_phone = form.get("supplier_phone") or None
    cash_from_drawer = float(form.get("cash_from_drawer", 0) or 0)
    capital_from_owner = float(form.get("capital_from_owner", 0) or 0)
    notes = form.get("notes") or None

    names = form.getlist("item_name[]")
    models = form.getlist("item_model[]")
    qtys = form.getlist("item_qty[]")
    prices = form.getlist("item_price[]")

    items = []
    for i, name in enumerate(names):
        if name.strip():
            qty = int(qtys[i]) if i < len(qtys) and qtys[i] else 1
            price = float(prices[i]) if i < len(prices) and prices[i] else 0.0
            items.append({
                "product_name": name.strip(),
                "model": models[i] if i < len(models) else None,
                "quantity": qty,
                "unit_price": price,
            })

    if not items:
        return RedirectResponse(f"/purchase-invoices/add?error={_q('أضف صنفاً واحداً على الأقل')}", status_code=302)

    total = sum(i["unit_price"] * i["quantity"] for i in items)
    if cash_from_drawer + capital_from_owner == 0:
        cash_from_drawer = total

    result = await api("post", "/api/purchase-invoices/", token=_token(request), json={
        "supplier_name": supplier_name,
        "supplier_phone": supplier_phone,
        "cash_from_drawer": cash_from_drawer,
        "capital_from_owner": capital_from_owner,
        "notes": notes,
        "items": items,
    })
    if isinstance(result, dict) and result.get("invoice_number"):
        inv_id = result.get("id")
        return RedirectResponse(f"/purchase-invoices/{inv_id}?success={_q(result.get('message','تم الحفظ'))}", status_code=302)
    err = result.get("detail", "فشل الحفظ") if isinstance(result, dict) else "فشل الحفظ"
    return RedirectResponse(f"/purchase-invoices/add?error={_q(err)}", status_code=302)


@app.get("/purchase-invoices/{invoice_id}")
async def purchase_invoice_detail(request: Request, invoice_id: int):
    if not _logged(request):
        return _redirect_login()
    inv = to_obj(await api("get", f"/api/purchase-invoices/{invoice_id}", token=_token(request)) or {})
    return templates.TemplateResponse(request, "purchase_invoice_detail.html", {
        "admin_name": _name(request),
        "active": "purchase-invoices",
        "invoice": inv,
        "success": request.query_params.get("success"),
        "error": request.query_params.get("error"),
    })


@app.post("/purchase-invoices/{invoice_id}/print")
async def purchase_invoice_mark_print(request: Request, invoice_id: int):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("put", f"/api/purchase-invoices/{invoice_id}/mark-printed", token=_token(request))
    if err:
        return RedirectResponse(f"/purchase-invoices/{invoice_id}?error={_q(err)}", status_code=302)
    return RedirectResponse(f"/purchase-invoices/{invoice_id}?success={_q('تم تسجيل الطباعة بنجاح')}", status_code=302)


# ══════════════════════════════════════════════════════════════════════════════
#  GALLERY
# ══════════════════════════════════════════════════════════════════════════════
@app.get("/gallery")
async def gallery_list(request: Request):
    if not _logged(request):
        return _redirect_login()
    series_list = to_obj(await api("get", "/api/gallery/folders", token=_token(request)) or [])
    return templates.TemplateResponse(request, "gallery.html", {
        "admin_name": _name(request),
        "active": "gallery",
        "series_list": series_list,
        "success": request.query_params.get("success"),
        "error": request.query_params.get("error"),
    })


@app.get("/gallery/{folder_id}")
async def gallery_folder_detail(request: Request, folder_id: int):
    if not _logged(request):
        return _redirect_login()
    folder = to_obj(await api("get", f"/api/gallery/folders/{folder_id}", token=_token(request)) or {})
    series_list = to_obj(await api("get", "/api/gallery/folders", token=_token(request)) or [])
    return templates.TemplateResponse(request, "gallery_folder.html", {
        "admin_name": _name(request),
        "active": "gallery",
        "folder": folder,
        "series_list": series_list,
        "success": request.query_params.get("success"),
        "error": request.query_params.get("error"),
    })


@app.post("/gallery/{folder_id}/upload")
async def gallery_upload_image(
    request: Request,
    folder_id: int,
    file: UploadFile = File(...),
    title: Optional[str] = Form(None),
):
    if not _logged(request):
        return _redirect_login()
    import io
    content = await file.read()
    files_payload = {"file": (file.filename, io.BytesIO(content), file.content_type)}
    data = {}
    if title:
        data["title"] = title
    result, err = await api_raw_upload(f"/api/gallery/folders/{folder_id}/images",
                                       files_payload, data, token=_token(request))
    if err:
        return RedirectResponse(f"/gallery/{folder_id}?error={_q(err)}", status_code=302)
    return RedirectResponse(f"/gallery/{folder_id}?success={_q('تم رفع الصورة')}", status_code=302)


@app.post("/gallery/images/{image_id}/delete")
async def gallery_delete_image(request: Request, image_id: int, folder_id: int = Form(...)):
    if not _logged(request):
        return _redirect_login()
    _, err = await api_ex("delete", f"/api/gallery/images/{image_id}", token=_token(request))
    if err:
        return RedirectResponse(f"/gallery/{folder_id}?error={_q(err)}", status_code=302)
    return RedirectResponse(f"/gallery/{folder_id}?success={_q('تم حذف الصورة بنجاح')}", status_code=302)


# ══════════════════════════════════════════════════════════════════════════════
#  EXPORT / INVENTORY  (original, keep below)
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/export/inventory")
async def export_inventory_excel(request: Request):
    if not _logged(request):
        return _redirect_login()
    raw = await api("get", "/api/inventory/", token=_token(request), params={"limit": 5000}) or []
    wb, ws = _make_wb("المخزون")
    _style_header(ws, ["#", "الرقم التسلسلي", "الفئة", "الماركة", "الموديل", "الدرجة", "السعر", "الحالة", "الفرع", "تاريخ الإضافة"])
    _set_widths(ws, [6, 22, 15, 15, 18, 10, 12, 12, 18, 20])
    status_labels = {"available":"متوفر","reserved":"محجوز","sold":"مباع"}
    for i, item in enumerate(raw, 1):
        created = (item.get("created_at") or "")[:16].replace("T", " ")
        branch = item.get("branch") or {}
        ws.append([i, item.get("serial_number",""), item.get("category",""),
                   item.get("brand",""), item.get("model",""), item.get("grade",""),
                   item.get("price",0),
                   status_labels.get(item.get("status",""), item.get("status","")),
                   branch.get("name","") if isinstance(branch, dict) else "",
                   created])
    return _wb_response(wb, f"inventory_{_date.today()}.xlsx")


# ── Settings ─────────────────────────────────────────────────────────────────

@app.get("/settings", response_class=HTMLResponse)
async def settings_page(request: Request):
    if not _logged(request):
        return _redirect_login()
    import os as _os
    data = await api("get", "/api/settings/", token=_token(request)) or {}
    sms_api_key_db = data.get("sms_api_key", "")
    _DEFAULT_KEY   = "37eaa347c97fb746d46eaf3d8fdb41737eeec5df"
    _DEFAULT_PHONE = "+967774440982"
    sms_api_key_env = _os.getenv("SMS_API_KEY", "") or _DEFAULT_KEY
    sms_api_key = sms_api_key_db or sms_api_key_env
    sms_key_via_env = bool(sms_api_key_env) and not sms_api_key_db
    sms_devices  = data.get("sms_devices", "0")
    sms_test_phone = _os.getenv("SMS_TEST_PHONE", "") or _DEFAULT_PHONE
    return templates.TemplateResponse(request, "settings.html", {
        "admin_name":      _name(request),
        "active":          "settings",
        "sms_configured":  bool(sms_api_key),
        "sms_key_via_env": sms_key_via_env,
        "sms_api_key":     sms_api_key,
        "sms_devices":     sms_devices,
        "sms_test_phone":  sms_test_phone,
    })


@app.post("/settings/sms", response_class=HTMLResponse)
async def settings_sms_save(
    request: Request,
    sms_api_key: str = Form(""),
    sms_devices:  str = Form("0"),
):
    if not _logged(request):
        return _redirect_login()
    token = _token(request)
    errors = []
    if sms_api_key.strip():
        _, err = await api_ex("post", "/api/settings/", token=token,
                              json={"key": "sms_api_key", "value": sms_api_key.strip()})
        if err:
            errors.append(err)
    _, err = await api_ex("post", "/api/settings/", token=token,
                          json={"key": "sms_devices", "value": sms_devices.strip() or "0"})
    if err:
        errors.append(err)
    if errors:
        from urllib.parse import quote as _q
        return RedirectResponse(f"/settings?error={_q('; '.join(errors))}", status_code=302)
    return RedirectResponse("/settings?success=تم+حفظ+إعدادات+SMS+بنجاح", status_code=302)


@app.post("/settings/sms/clear")
async def settings_sms_clear(request: Request):
    if not _logged(request):
        return _redirect_login()
    token = _token(request)
    await api_ex("delete", "/api/settings/sms_api_key", token=token)
    await api_ex("delete", "/api/settings/sms_devices",  token=token)
    return RedirectResponse("/settings?success=تم+حذف+مفتاح+SMS", status_code=302)


@app.post("/settings/sms/test")
async def settings_sms_test(request: Request):
    """AJAX endpoint — returns JSON {ok, message}"""
    from fastapi.responses import JSONResponse
    if not _logged(request):
        return JSONResponse({"ok": False, "message": "غير مصرَّح — سجّل دخولك"}, status_code=401)
    try:
        body = await request.json()
        phone = (body.get("phone") or "").strip()
    except Exception:
        return JSONResponse({"ok": False, "message": "رقم الجوال مطلوب"}, status_code=400)
    if not phone:
        return JSONResponse({"ok": False, "message": "رقم الجوال مطلوب"}, status_code=400)

    _, err = await api_ex("post", "/api/auth/send-otp", token=_token(request),
                          json={"phone": phone, "resend": True})
    if err:
        return JSONResponse({"ok": False, "message": err})
    return JSONResponse({"ok": True, "message": f"تم إرسال رسالة تجريبية إلى {phone}"})
